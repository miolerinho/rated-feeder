"""
Bank of England OIS Rate Fetcher

Fetches OIS (Overnight Index Swap) rates from the Bank of England's daily data
and extracts 2yr, 5yr, and 10yr rates with day-over-day changes.
"""

import requests
from zipfile import ZipFile
from io import BytesIO
import openpyxl
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple
import warnings

warnings.filterwarnings('ignore', category=UserWarning)


class BOEOISFetcher:
    """Fetches and parses Bank of England OIS rate data."""

    BOE_URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
    OIS_FILENAME = "OIS daily data current month.xlsx"
    SHEET_NAME = "4. spot curve"

    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }

    def fetch_ois_data(self) -> Tuple[Optional[Dict], Optional[str]]:
        """
        Fetch OIS data from Bank of England.

        Returns:
            Tuple of (data_dict, error_message)
            data_dict contains latest and previous rates with changes
        """
        try:
            # Download ZIP file
            response = requests.get(self.BOE_URL, headers=self.headers, timeout=30)
            response.raise_for_status()

            # Extract OIS Excel file
            with ZipFile(BytesIO(response.content)) as zip_file:
                with zip_file.open(self.OIS_FILENAME) as ois_file:
                    wb = openpyxl.load_workbook(ois_file)
                    sheet = wb[self.SHEET_NAME]

                    # Parse the data
                    return self._parse_ois_sheet(sheet), None

        except requests.RequestException as e:
            return None, f"Network error: {str(e)}"
        except KeyError as e:
            return None, f"Sheet not found: {str(e)}"
        except Exception as e:
            return None, f"Error fetching OIS data: {str(e)}"

    def _parse_ois_sheet(self, sheet) -> Dict:
        """
        Parse OIS spot curve sheet to extract 2yr, 5yr, 10yr rates.

        Sheet structure:
        - Row 4 contains maturity years (0.5, 1, 1.5, 2, 2.5, 3, ..., 5, ..., 10, ...)
        - Row 6+ contains dates and rates
        """
        # Find column indices for 2yr, 5yr, 10yr
        maturity_row = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]

        col_2yr = None
        col_5yr = None
        col_10yr = None

        for idx, value in enumerate(maturity_row):
            if value == 2:
                col_2yr = idx
            elif value == 5:
                col_5yr = idx
            elif value == 10:
                col_10yr = idx

        if col_2yr is None or col_5yr is None or col_10yr is None:
            raise ValueError("Could not find 2yr, 5yr, or 10yr columns")

        # Extract latest and previous day's data
        data_rows = []
        for row in sheet.iter_rows(min_row=6, values_only=True):
            date_val = row[0]
            if isinstance(date_val, datetime):
                data_rows.append({
                    'date': date_val,
                    'rate_2yr': row[col_2yr],
                    'rate_5yr': row[col_5yr],
                    'rate_10yr': row[col_10yr]
                })

        # Sort by date descending
        data_rows.sort(key=lambda x: x['date'], reverse=True)

        if len(data_rows) < 2:
            raise ValueError("Insufficient data to calculate changes")

        latest = data_rows[0]
        previous = data_rows[1]

        # Calculate changes (in basis points)
        def calculate_change(latest_rate, prev_rate):
            if latest_rate is not None and prev_rate is not None:
                return (latest_rate - prev_rate) * 100  # Convert to basis points
            return None

        result = {
            'latest_date': latest['date'],
            'previous_date': previous['date'],
            'rates': {
                '2yr': {
                    'current': latest['rate_2yr'],
                    'previous': previous['rate_2yr'],
                    'change_bps': calculate_change(latest['rate_2yr'], previous['rate_2yr'])
                },
                '5yr': {
                    'current': latest['rate_5yr'],
                    'previous': previous['rate_5yr'],
                    'change_bps': calculate_change(latest['rate_5yr'], previous['rate_5yr'])
                },
                '10yr': {
                    'current': latest['rate_10yr'],
                    'previous': previous['rate_10yr'],
                    'change_bps': calculate_change(latest['rate_10yr'], previous['rate_10yr'])
                }
            }
        }

        return result

    def format_summary(self, data: Dict) -> str:
        """Format OIS data into a readable summary."""
        latest_date = data['latest_date'].strftime('%d %B %Y')
        previous_date = data['previous_date'].strftime('%d %B %Y')

        summary = f"""Bank of England OIS Rates Summary
{'='*50}

Latest Data: {latest_date}
Previous Data: {previous_date}

"""

        for tenor, values in data['rates'].items():
            current = values['current']
            previous = values['previous']
            change = values['change_bps']

            change_str = f"{change:+.2f} bps" if change is not None else "N/A"
            arrow = "↑" if change and change > 0 else "↓" if change and change < 0 else "→"

            summary += f"{tenor:>4} Rate: {current:>6.3f}% (was {previous:>6.3f}%) {arrow} {change_str}\n"

        summary += f"\n{'='*50}\n"
        summary += f"Data source: Bank of England\n"
        summary += f"Fetched at: {datetime.now().strftime('%d %B %Y %H:%M:%S')}\n"

        return summary


def main():
    """Test the OIS fetcher."""
    fetcher = BOEOISFetcher()

    print("Fetching Bank of England OIS data...")
    data, error = fetcher.fetch_ois_data()

    if error:
        print(f"Error: {error}")
        return

    if data:
        print(fetcher.format_summary(data))
    else:
        print("No data returned")


if __name__ == "__main__":
    main()